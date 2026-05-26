from PIL import Image, ImageDraw, ImageFont
import re
from dataclasses import dataclass
from collections import defaultdict
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook
from database import db


@dataclass
class Lesson:
    time_str: str
    subject_name: str = ""
    teacher_name: str = ""
    classroom: str = ""
    raw_text: str = ""


@dataclass
class GroupSchedule:
    group_number: str
    group_label: str
    direction: str
    week_label: str
    lessons: list[Lesson] = None


# ====================== РАБОТА С БД ======================


async def save_schedule_to_db(schedule: GroupSchedule):
    """Сохраняет / обновляет расписание на завтра"""
    async with db.pool.acquire() as conn:
        group_id = await conn.fetchval(
            'SELECT id_g FROM "Group" WHERE group_number = $1',
            int(schedule.group_number)
        )

        if not group_id:
            print(f"⚠️ Группа {schedule.group_number} не найдена")
            return

        # Удаляем старые записи на завтра (чистое обновление)
        await conn.execute(
            """
            DELETE FROM Schedule 
            WHERE id_group = $1 
              AND dates = CURRENT_DATE + INTERVAL '1 day'
            """,
            group_id
        )

        for lesson in schedule.lessons:
            start_time = None
            end_time = None
            if '-' in lesson.time_str:
                parts = [p.strip() for p in lesson.time_str.split('-')]
                try:
                    start_time = datetime.strptime(parts[0], "%H:%M").time()
                    if len(parts) > 1:
                        end_time = datetime.strptime(parts[1], "%H:%M").time()
                except ValueError:
                    pass

            subject_name = (lesson.subject_name or "Не указано")[:180]
            classroom_num = (lesson.classroom or "—")[:45]

            # Предмет
            subject_id = await conn.fetchval(
                """
                INSERT INTO Subject (name) 
                VALUES ($1) 
                ON CONFLICT (name) DO UPDATE SET name = EXCLUDED.name 
                RETURNING id_sub
                """,
                subject_name
            )

            # Аудитория
            classroom_id = await conn.fetchval(
                """
                INSERT INTO Classroom (number) 
                VALUES ($1) 
                ON CONFLICT (number) DO UPDATE SET number = EXCLUDED.number 
                RETURNING id_cs
                """,
                classroom_num
            )

            # === Вставка / обновление расписания ===
            await conn.execute(
                """
                INSERT INTO Schedule 
                (dates, start_time, end_time, id_group, id_subject, id_classroom, week_label)
                VALUES (CURRENT_DATE + INTERVAL '1 day', $1, $2, $3, $4, $5, $6)
                ON CONFLICT DO NOTHING
                """,
                start_time,
                end_time,
                group_id,
                subject_id,
                classroom_id,
                schedule.week_label
            )

        print(f"✅ Обновлено на завтра: {schedule.group_label} | {len(schedule.lessons)} пар")


async def get_group_schedule_from_db(group_number: str, week_label: str = None) -> list[Lesson] | None:
    async with db.pool.acquire() as conn:
        group_id = await conn.fetchval('SELECT id_g FROM "Group" WHERE group_number = $1', int(group_number))
        if not group_id:
            return None

        rows = await conn.fetch(
            """
            SELECT s.start_time, s.end_time, sub.name AS subject_name,
                   c.number AS classroom, COALESCE(u.fam || ' ' || u.name, '') AS teacher_name
            FROM Schedule s
            LEFT JOIN Subject sub ON s.id_subject = sub.id_sub
            LEFT JOIN Classroom c ON s.id_classroom = c.id_cs
            LEFT JOIN Teacher t ON s.id_teacher = t.id_t
            LEFT JOIN Users u ON t.id_user = u.id_us
            WHERE s.id_group = $1 AND ($2::text IS NULL OR s.week_label = $2)
            ORDER BY s.start_time
            """,
            group_id, week_label
        )

        return [Lesson(
            time_str=f"{r['start_time']}–{r['end_time']}" if r['end_time'] else str(r['start_time']),
            subject_name=r['subject_name'] or "",
            teacher_name=r['teacher_name'],
            classroom=r['classroom'] or ""
        ) for r in rows]


# ====================== ПАРСИНГ EXCEL ======================

def parse_lesson_text(raw_text: str) -> Lesson:
    """Улучшенный разбор текста занятия"""
    if not raw_text or str(raw_text).strip() in ("", "None", "nan"):
        return Lesson(time_str="", subject_name="", raw_text="")

    text = str(raw_text).strip()
    lines = [line.strip() for line in text.split('\n') if line.strip()]

    subject = lines[0] if lines else ""
    teacher = ""
    classroom = ""

    for line in lines[1:]:
        if re.search(r'\d-\d', line) or any(x in line.lower() for x in ["ауд", "каб", "корп", "этаж"]):
            classroom = line
        elif line and len(line) > 3:
            teacher = line

    return Lesson(
        time_str="",
        subject_name=subject,
        teacher_name=teacher,
        classroom=classroom,
        raw_text=text
    )


def parse_excel(file_path: str) -> list[GroupSchedule]:
    """Адаптировано под структуру твоего файла"""
    wb = load_workbook(file_path, data_only=True)
    result = []

    for sheet in wb.worksheets:
        sheet_name = sheet.title
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 3:
            continue

        row_directions = rows[0]
        row_groups = rows[1]
        pair_rows = rows[2:]

        groups = {}
        last_direction = ""

        for col_idx in range(1, len(row_groups)):
            # Определяем направление
            if col_idx < len(row_directions) and row_directions[col_idx]:
                last_direction = str(row_directions[col_idx]).strip()

            raw_group = row_groups[col_idx]
            if not raw_group:
                continue

            label = str(raw_group).strip()
            m = re.search(r'(\d+)', label)
            if not m:
                continue

            gs = GroupSchedule(
                group_number=m.group(1),
                group_label=label,
                direction=last_direction,
                week_label=sheet_name,
                lessons=[]
            )
            groups[col_idx] = gs
            result.append(gs)

        # Заполняем пары
        for row in pair_rows:
            if not row or not row[0]:
                continue

            time_cell = str(row[0]).strip()
            # Убираем лишние переносы в номере пары
            time_cell = re.sub(r'^\d+\s*', '', time_cell).strip()

            if not time_cell or time_cell.lower() in ("none", ""):
                continue

            for col_idx, gs in groups.items():
                cell_val = row[col_idx] if col_idx < len(row) else None
                text = str(cell_val).strip() if cell_val is not None else ""

                lesson = parse_lesson_text(text)
                lesson.time_str = time_cell
                gs.lessons.append(lesson)

    return result


async def process_and_save_excel(file_path: str, context=None):
    """Полная обработка Excel + уведомления"""
    print(f"📂 Читаем файл: {file_path}")
    schedules = parse_excel(file_path)

    print(f"Найдено групп: {len(schedules)}")

    updated_group_ids = []

    for gs in schedules:
        group_id = await get_group_id(gs.group_number)  # вспомогательная функция
        if group_id:
            await save_schedule_to_db(gs)
            updated_group_ids.append(group_id)

    print("🎉 Загрузка расписания завершена!")

    # Отправляем уведомления
    if updated_group_ids and context:
        await notify_groups_about_update(updated_group_ids, context)

    return schedules


# Вспомогательная функция
async def get_group_id(group_number: str) -> int | None:
    async with db.pool.acquire() as conn:
        return await conn.fetchval(
            'SELECT id_g FROM "Group" WHERE group_number = $1',
            int(group_number)
        )


# ====================== ГЕНЕРАЦИЯ ИЗОБРАЖЕНИЯ ======================

COLOR_HEADER_BG = (224, 100, 100)
COLOR_HEADER_TEXT = (255, 255, 255)
COLOR_TIME_BG = (245, 245, 245)
COLOR_TIME_TEXT = (60, 60, 60)
COLOR_CELL_BG = (255, 255, 255)
COLOR_CELL_TEXT = (30, 30, 30)
COLOR_BORDER = (190, 190, 190)
COLOR_WEEK_BG = (210, 70, 70)

FONT_SIZE_HEADER = 36
FONT_SIZE_WEEK = 27
FONT_SIZE_TIME = 22
FONT_SIZE_CELL = 20

COL_TIME_W = 195        # Увеличил ширину времени
COL_LESSON_W = 485
PADDING = 14
IMG_WIDTH = COL_TIME_W + COL_LESSON_W


def _font(size: int, bold: bool = False):
    """Старый стиль шрифтов (DejaVu / Liberation)"""
    candidates_bold = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        "C:/Windows/Fonts/arialbd.ttf",
    ]
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "C:/Windows/Fonts/arial.ttf",
    ]
    for path in (candidates_bold if bold else candidates):
        if Path(path).exists():
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


def _wrap(text: str, font, max_w: int) -> list[str]:
    if not text:
        return [""]
    result = []
    for paragraph in text.split('\n'):
        words = paragraph.split()
        line = ""
        for word in words:
            test = (line + " " + word).strip()
            if font.getbbox(test)[2] <= max_w:
                line = test
            else:
                if line:
                    result.append(line)
                line = word
        if line:
            result.append(line)
    return result or [""]


def _format_time_range(time_str: str) -> str:
    """Приводит время к формату '08:40–10:15' без секунд"""
    if not time_str:
        return ""

    # Если это уже строка с диапазоном
    if '–' in time_str:
        parts = time_str.split('–')
    elif '-' in time_str:
        parts = time_str.split('-')
    else:
        return time_str[:5]  # одиночное время

    cleaned = []
    for part in parts:
        part = part.strip()
        if ':' in part:
            # Оставляем только HH:MM
            cleaned.append(part[:5])
        else:
            cleaned.append(part)

    return '–'.join(cleaned)


def render_group_image(gs: GroupSchedule, output_path: str) -> str:
    """Заголовок: 'Расписание для группы XXXX'"""
    f_header = _font(FONT_SIZE_HEADER, bold=True)
    f_week = _font(FONT_SIZE_WEEK, bold=True)
    f_time = _font(FONT_SIZE_TIME, bold=True)
    f_cell = _font(FONT_SIZE_CELL)

    tomorrow = datetime.now() + timedelta(days=1)
    date_str = tomorrow.strftime("%d.%m.%Y") + " (Завтра)"

    # Заголовок с номером группы
    group_title = f"Расписание для группы {gs.group_number}"

    # Расчёт высоты строк
    row_heights = []
    for lesson in gs.lessons:
        time_str = _format_time_range(lesson.time_str)
        time_h = len(_wrap(time_str, f_time, COL_TIME_W - PADDING * 2)) * 29 + PADDING * 2

        lesson_text = f"{lesson.subject_name}\n{lesson.teacher_name}\n{lesson.classroom}".strip()
        lesson_h = len(_wrap(lesson_text, f_cell, COL_LESSON_W - PADDING * 2)) * 26 + PADDING * 2

        row_heights.append(max(88, time_h, lesson_h))

    H_TOP = 95
    H_WEEK = 58
    total_h = H_TOP + H_WEEK + sum(row_heights) + 30

    img = Image.new("RGB", (IMG_WIDTH, total_h), COLOR_CELL_BG)
    draw = ImageDraw.Draw(img)
    y = 0

    # Главный заголовок с группой
    draw.rectangle([0, y, IMG_WIDTH, y + H_TOP], fill=COLOR_HEADER_BG)
    title_w = f_header.getbbox(group_title)[2]
    draw.text(((IMG_WIDTH - title_w) // 2, y + 26), group_title, font=f_header, fill=COLOR_HEADER_TEXT)
    y += H_TOP

    # Дата
    draw.rectangle([0, y, IMG_WIDTH, y + H_WEEK], fill=COLOR_WEEK_BG)
    date_w = f_week.getbbox(date_str)[2]
    draw.text(((IMG_WIDTH - date_w) // 2, y + 14), date_str, font=f_week, fill=COLOR_HEADER_TEXT)
    y += H_WEEK

    # Пары
    for lesson, row_h in zip(gs.lessons, row_heights):
        draw.rectangle([0, y, IMG_WIDTH, y + row_h], fill=COLOR_CELL_BG)
        draw.rectangle([0, y, COL_TIME_W, y + row_h], fill=COLOR_TIME_BG)

        # Время начала — конец
        time_str = _format_time_range(lesson.time_str)
        time_lines = _wrap(time_str, f_time, COL_TIME_W - PADDING * 2)
        for i, line in enumerate(time_lines):
            draw.text((PADDING + 4, y + PADDING + i * 29), line, font=f_time, fill=COLOR_TIME_TEXT)

        # Предмет + преподаватель + кабинет
        lesson_text = lesson.subject_name.strip()
        if lesson.teacher_name:
            lesson_text += f"\n{lesson.teacher_name}"
        if lesson.classroom:
            lesson_text += f"\n{lesson.classroom}"

        lesson_lines = _wrap(lesson_text, f_cell, COL_LESSON_W - PADDING * 2)
        for i, line in enumerate(lesson_lines):
            draw.text((COL_TIME_W + PADDING, y + PADDING + i * 26), line, font=f_cell, fill=COLOR_CELL_TEXT)

        # Границы
        draw.line([(0, y + row_h), (IMG_WIDTH, y + row_h)], fill=COLOR_BORDER, width=1)
        draw.line([(COL_TIME_W, y), (COL_TIME_W, y + row_h)], fill=COLOR_BORDER, width=2)

        y += row_h

    # Внешняя рамка
    draw.rectangle([0, 0, IMG_WIDTH - 1, total_h - 1], outline=(160, 160, 160), width=4)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    img.save(output_path, "PNG")
    print(f"🖼️ Изображение сохранено: {output_path}")
    return output_path

async def get_tomorrow_schedule_for_user(telegram_id: int) -> list[Lesson] | None:
    """Возвращает расписание на завтра для пользователя"""
    async with db.pool.acquire() as conn:
        # Получаем группу пользователя
        group_id = await conn.fetchval(
            'SELECT id_group FROM Users WHERE telegram_id = $1',
            telegram_id
        )
        if not group_id:
            return None

        # Завтрашний день
        rows = await conn.fetch(
            """
            SELECT 
                s.start_time, s.end_time, 
                sub.name AS subject_name,
                c.number AS classroom,
                COALESCE(u.fam || ' ' || u.name, '') AS teacher_name
            FROM Schedule s
            LEFT JOIN Subject sub ON s.id_subject = sub.id_sub
            LEFT JOIN Classroom c ON s.id_classroom = c.id_cs
            LEFT JOIN Teacher t ON s.id_teacher = t.id_t
            LEFT JOIN Users u ON t.id_user = u.id_us
            WHERE s.id_group = $1 
              AND s.dates = CURRENT_DATE + INTERVAL '1 day'
            ORDER BY s.start_time
            """,
            group_id
        )

        return [Lesson(
            time_str=f"{r['start_time']}–{r['end_time']}" if r['end_time'] else str(r['start_time']),
            subject_name=r['subject_name'] or "",
            teacher_name=r['teacher_name'],
            classroom=r['classroom'] or ""
        ) for r in rows]

# ====================== УВЕДОМЛЕНИЯ ======================

async def notify_groups_about_update(updated_groups: list[int], context=None):
    """Отправляет уведомление всем пользователям обновлённых групп"""
    if not context or not updated_groups:
        return

    async with db.pool.acquire() as conn:
        users = await conn.fetch(
            """
            SELECT telegram_id, id_group 
            FROM Users 
            WHERE id_group = ANY($1) AND telegram_id IS NOT NULL
            """,
            updated_groups
        )

    if not users:
        print("ℹ️ Никто не зарегистрирован в обновлённых группах")
        return

    success_count = 0
    for user in users:
        try:
            await context.bot.send_message(
                chat_id=user['telegram_id'],
                text="🔔 Расписание на завтра обновлено!\n\n"
                     "Используй команду /schedule, чтобы посмотреть актуальное расписание.",
                parse_mode='HTML'
            )
            success_count += 1
        except Exception as e:
            print(f"Не удалось отправить уведомление пользователю {user['telegram_id']}: {e}")

    print(f"✅ Уведомления отправлены {success_count} пользователям из {len(updated_groups)} групп")
